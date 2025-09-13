import pandas as pd

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# === Configuration ===
input_file = r"C:\Users\Dell\Desktop\Python Practice\source_datafiles\excelsheets\German_Verbs.xlsx"   # Input Excel file with verbs
output_folder = r"C:\Users\Dell\Desktop\Python Practice\outputfiles\cleaned_data"
output_file = os.path.join(output_folder, "Processed_Verbs.xlsx")

# Create output folder if it does not exist
os.makedirs(output_folder, exist_ok=True)

# === Load Excel file ===
df = pd.read_excel(input_file)

# === Find duplicates in column "Verb" ===
duplicates = df[df.duplicated(subset=["Verb"], keep=False)]
unique_verbs = df.drop_duplicates(subset=["Verb"], keep="first")

# === Save results into one Excel file with two sheets ===
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    unique_verbs.to_excel(writer, sheet_name="Unique_Verbs", index=False)
    duplicates.to_excel(writer, sheet_name="Duplicate_Verbs", index=False)

# === Add colors and formatting ===
wb = load_workbook(output_file)

# Define fills
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue header
unique_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
dup_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red

# Style both sheets
for sheet_name in ["Unique_Verbs", "Duplicate_Verbs"]:
    ws = wb[sheet_name]
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)  # White bold text
    
    # Style rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if sheet_name == "Unique_Verbs":
                cell.fill = unique_fill
            else:
                cell.fill = dup_fill

# Save styled workbooks
wb.save(output_file)

print("✅ Done! Colorful Excel file saved in:", output_file)